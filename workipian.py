##
##  @Author: Sonu Gupta
##  @Date: 3-2-18
##  @purpose: [Personal Utility tool] GUI application which adds the 'Tasks' done to the excel worksheet.
##            Often I keep note of the tasks done on day-today basis.
##            This helps me during "appraisal time". :D Because I can always look back at the work I have done which I have noted in excelsheet.
##            Using this excelsheet i can make sure I am not missing anything for final consolidation of thw work.
##
##
##  Note: prerequisite: modules--> openpyxl, tkinter

## todo: Content checking

import openpyxl
import os
from datetime import datetime
from openpyxl.styles import Alignment
from tkinter import *
import tkinter.messagebox

# Workbook name
filename = "my_Work_sheet.xlsx"
sheetname= "Sheet1"

# Function which updates the workbook at backend.
def updateWorksheet(data):
    wb = openpyxl.load_workbook(filename)
    print(wb.get_sheet_names())

    sheet = wb.get_sheet_by_name(sheetname)

    # Column 'A' always depicts the todays date.
    today = datetime.strftime(datetime.now(), ' %d-%m-%y')
    sheet.append([today, data])

    # This gives you the active worksheet.
    ws = wb.active

    # This is done so that you dont need to adjust column length manually everytime.
    ws.column_dimensions["B"].width = 60.0

    # By default, wrap_text is false. This is done so that you can see wrapped text neatly.
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)
    wb.save(filename)


def addRecord():

    # Reading from text box:
    # The first part, "1.0" means that the input should be read from line one, character zero.
    # END is an imported constant which is set to the string "end". The END part means to read until the end of the text box is reached.
    # The only issue with this is that it actually adds a newline to our input.
    # So, in order to fix it we should change END to end-1c.
    #  The -1c deletes 1 character, while -2c would mean delete two characters, and so on.
    try:
        data = editBox.get("1.0","end-1c")
        updateWorksheet(data)
        print("Record Added succefully..!!")
        tkinter.messagebox.showinfo('Success', 'Work Record is added successfully !!!')

        # disabling the 'add' button. User can add only one data per day. [Did this for simplicity.]
        addButton.config(state='disabled')
    except PermissionError:
        tkinter.messagebox.showerror('Error', 'Excelsheet is already open. Please close it.')


# checking if Workbook exists or not.

if True != os.path.exists(filename):
    # Create workbook if not exists
    wb = openpyxl.Workbook() # This creates new workbook
    wb.create_sheet(index=0, title=sheetname)          # by default added to last
    wb.save(filename)
    print('No such file exists created')

# main routine
root = Tk()
root.title("Personal Worksheet")
root.resizable(0,0)
root.iconbitmap('writer.ico')

frame = Frame(root, width=500, height=500)

editBox= Text(frame, height=10, width=60)
editBox.pack(pady=20)

addButton = Button(frame, text="Add", command=addRecord, width=10, height=2)
addButton.pack(pady=10)
frame.pack()

root.mainloop()
